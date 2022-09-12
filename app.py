from flask import Flask, render_template
import openpyxl

app = Flask(__name__)


def data_colloct_from_excel(excel_name):
    work_book = openpyxl.load_workbook(excel_name)
    sheet_names_all = work_book.sheetnames
    list_all = []
    for k in range(len(sheet_names_all)):
        sheet_name = sheet_names_all[k]
        sheet = work_book[sheet_name]
        row = sheet.max_row
        column = sheet.max_column
        list1 = []
        for i in range(row):
            row_list = []
            for j in range(column):
                row_list.append((sheet.cell(i + 1, j + 1)).value)
            list1.append(row_list)
        list_all.append(list1)
    return list_all


@app.route('/')
def hello_world():  # put application's code here
    list_all = (data_colloct_from_excel("province_in_diff_days.xlsx"))[0]
    return render_template("test.html", list_all=list_all)


if __name__ == '__main__':
    app.run()
