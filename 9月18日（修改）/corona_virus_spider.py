import time
import requests
import re
import openpyxl

from bs4 import BeautifulSoup
from openpyxl.styles import Alignment
from tqdm import tqdm


# 爬取并生成每天的疫情数据，并存入excel表，每次爬取一个网页睡眠1秒
class CoronaVirusSpider(object):
    # 初始化方法
    def __init__(self, new_excel_name='new1'):
        self.new_excel_name = new_excel_name + '.xlsx'
        self.home_url = 'http://www.nhc.gov.cn/xcs/yqtb/list_gzbd.shtml'
        self.home_header = {
            "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,*/*;q=0.8",
            "Accept-Encoding": "gzip, deflate",
            "Accept-Language": "zh-CN,zh;q=0.8,zh-TW;q=0.7,zh-HK;q=0.5,en-US;q=0.3,en;q=0.2",
            "Connection": "keep-alive",
            "Referer": "http://www.nhc.gov.cn/",
            "Cookie": "yfx_c_g_u_id_10006654=_ck22090517092711557123132504572; yfx_f_l_v_t_10006654=f_t_1662368967158_\
            _r_t_1662514638127__v_t_1662514638127__r_c_2; yfx_mr_10006654=%3A%3Amarket_type_free_search%3A%3A%3A%3A\
            baidu%3A%3A%3A%3A%3A%3A%3A%3Awww.baidu.com%3A%3A%3A%3Apmf_from_free_search; yfx_mr_f_10006654=%3A%3A\
            market_type_free_search%3A%3A%3A%3Abaidu%3A%3A%3A%3A%3A%3A%3A%3Awww.baidu.com%3A%3A%3A%3Apmf_from_free_search;\
             yfx_key_10006654=; sVoELocvxVW0S=57yh5eHi6BlWwbYuOEUHFMNXf_2SF8UL5VWS1759zdOiiImwtyLuvBL1rWffIpGlLnMEMoxnp\
             QBHoAej5Qug.gG; security_session_verify=fc465e4f1828940ec0438b63374ada0a; sVoELocvxVW0T=53SI0.DWUeQ7qqqDkm\
             RH3_AToYARjKiHRH568jKOM4B.OPNB2axXw5kqAtweBhHBYQOYh3hRO8OaMl8SZuRBb4HDDy8wWx_H9KnDMJfOHJhLKqwvylr_gmnhMbVf\
             7Xl1INInmRUZl8aTrrguv1MWZmyUOXCgg2aOx6_4J72Gm.uCLEdwxtjF7hWLGGpO..CyBUuKFNGN8o.f7i5cTf3DueMgKy959yMbnxH14vn\
             DsH.wVdK4nQbz4PLAMqYCxYgwjT4eY2xujVIScYsPVxnC5uNL45UyizBvMCagu5cjSPGfoWPa5mHqCrzryZOZ96c0axMDYQTxAdJ7Le\
             VECT_l6vDQTiBIU5G26AzdTJ07AoQza; insert_cookie=91349450",
            "Hosts": "www.nhc.gov.cn",
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:104.0) Gecko/20100101 Firefox/104.0"
        }
        self.province_name = [
            '安徽', '北京', '重庆', '福建', '甘肃', '广东', '广西', '贵州', '海南', '河北', '河南',
            '黑龙江', '湖北', '湖南', '吉林', '江苏', '江西', '辽宁', '内蒙古', '宁夏', '青海', '山东',
            '山西', '陕西', '上海', '四川', '天津', '西藏', '新疆', '云南', '浙江'
        ]

    # 删除初始化sheet
    def del_sheet1(self):
        work_book = openpyxl.load_workbook(self.new_excel_name)
        sheet_name = 'Sheet'
        worksheet = work_book[sheet_name]
        work_book.remove(worksheet)
        work_book.save(self.new_excel_name)

    # 将某天的数据字典信息存入excel表中，并修改excel表的格式
    def write_one_day_excel(self, new_day_name, nums_all, dict_1, dict_2):
        work_book = openpyxl.load_workbook(self.new_excel_name)
        sheet = work_book.create_sheet(new_day_name)
        arrli = self.province_name
        list_len = len(arrli)

        for i in range(list_len):
            province_name = arrli[i]
            sheet.cell(i + 3, 1, arrli[i])
            if province_name in dict_1:
                sheet.cell(i + 3, 2, int(dict_1[province_name]))
            else:
                sheet.cell(i + 3, 2, 0)
            if province_name in dict_2:
                sheet.cell(i + 3, 3, int(dict_2[province_name]))
            else:
                sheet.cell(i + 3, 3, 0)
        sheet.cell(1, 1, new_day_name)
        sheet.cell(1, 2, '本土新增确诊')
        sheet.cell(1, 3, '本土新增无症状')
        sheet.cell(2, 1, '中国内地')
        sheet.cell(2, 2, int(nums_all[0]))
        sheet.cell(2, 3, int(nums_all[1]))
        sheet.cell(list_len + 3, 1, '港澳台地区')
        sheet.cell(list_len + 3, 2, '累计通报确诊病例')
        sheet.cell(list_len + 4, 1, '香港')
        sheet.cell(list_len + 5, 1, '澳门')
        sheet.cell(list_len + 6, 1, '台湾')
        sheet.cell(list_len + 4, 2, int(nums_all[2]))
        sheet.cell(list_len + 5, 2, int(nums_all[3]))
        sheet.cell(list_len + 6, 2, int(nums_all[4]))

        for j in range(list_len + 7):
            for k in range(3):
                sheet.cell(row=j + 1, column=k + 1).alignment = Alignment(horizontal='center', vertical='center')
        sheet.column_dimensions['A'].width = 20.0
        sheet.column_dimensions['B'].width = 20.0
        sheet.column_dimensions['C'].width = 20.0

        work_book.save(self.new_excel_name)
        work_book.close()

    # 将字符串和数字用正则分离并存如字典
    @staticmethod
    def change_str_to_dict(str_main):
        str_key = re.findall(r'([\u4e00-\u9fa5]+)\d', str_main)
        str_value = re.findall(r'\d+', str_main)
        return dict(zip(str_key, str_value))

    # 处理某一天的网页爬虫信息，做数据分割
    def one_day_page(self, page_new, new_day):
        pattern = re.compile(r'<[^>]+>', re.S)
        result = str(pattern.sub('', page_new))
        # print(result)
        num_1 = re.findall(r'新增确诊病例.+本土病例(\d+)例', result)[0]
        data_1 = re.findall(r'新增确诊病例.+本土病例\d+例（(.+)），', result)[0]
        dict_1 = self.change_str_to_dict(data_1)

        num_2 = re.findall(r'新增无症状感染者.+本土(\d+)例', result)[0]
        data_2 = re.findall(r'新增无症状感染者.+本土\d+例（(.+)）。', result)[0]
        dict_2 = self.change_str_to_dict(data_2)

        num_3 = re.findall(r'香港特别行政区(\d+)例', result)[0]
        num_4 = re.findall(r'澳门特别行政区(\d+)例', result)[0]
        num_5 = re.findall(r'台湾地区(\d+)例', result)[0]
        nums_all = [num_1, num_2, num_3, num_4, num_5]
        # print(nums_all)
        datas_all = [new_day, nums_all, dict_1, dict_2]
        # self.write_one_day_excel(new_day, nums_all, dict_1, dict_2)
        return datas_all

    # 控制爬虫采集和存储到excel操作的核心控件
    def control_every_child_page(self, soup, page_num):
        day_tag_all = soup.findAll('span', attrs={"class": "ml"})
        page_name = [span.get_text() for span in day_tag_all]

        a_tag_all = soup.findAll('a', title=re.compile('疫情最新情况'))
        for i in tqdm(range(len(a_tag_all)), '采集第' + str(page_num) + '页每日疫情数据'):
            time.sleep(1)
            text_a = a_tag_all[i].attrs
            # print(text_a)
            # print(result[i])
            text_a['href'] = 'http://www.nhc.gov.cn' + text_a['href']
            del text_a['target']
            text_a['title'] = page_name[i]
            page_new = self.get_content_from_url(text_a['href'], self.home_header)
            # test = self.crawl_one_day_url(text_a['href'])
            # print(test)
            # print(page_new)
            # print(text_a['title'])
            dates_all = self.one_day_page(page_new, page_name[i])
            self.write_one_day_excel(dates_all[0], dates_all[1], dates_all[2], dates_all[3])
            # self.create_main_db(page_name[i])
            # self.write_one_day_db(dates_all[0], dates_all[1], dates_all[2], dates_all[3])
        return page_name

    # 获得某网页的源代码信息
    @staticmethod
    def get_content_from_url(my_url, my_headers):
        response = requests.get(my_url, headers=my_headers)
        return response.content.decode()

    # 可选择的使用bs4解析源代码
    def crawl_one_day_url(self, home_url):
        home_page = self.get_content_from_url(home_url, self.home_header)
        soup = BeautifulSoup(home_page, 'html.parser')
        return soup

    # 官网有好几页的信息，控制这些页面自身的url
    @staticmethod
    def return_different_home_page_url(page_num):
        # for i in range(begin_page, end_page):
        if page_num == 1:
            url_home_page = 'http://www.nhc.gov.cn/xcs/yqtb/list_gzbd.shtml'
        else:
            url_home_page = 'http://www.nhc.gov.cn/xcs/yqtb/list_gzbd_' + str(page_num) + '.shtml'
        return url_home_page

    # 最外层的控制组件，用于遍历从第开始页到结束页的链接信息
    def control_range_home_url(self, begin_page_num, end_page_num):
        page_name_all = []
        for i in range(begin_page_num, end_page_num + 1):
            url_home = self.return_different_home_page_url(i)
            # return url_childen
            url_children = self.crawl_one_day_url(url_home)
            # print(url_children)
            page_name = self.control_every_child_page(url_children, i)
            page_name_all = page_name_all + page_name
        return page_name_all

    # 创建一个储存每一天网页数据信息的excel表
    def create_main_excel(self):
        work_book = openpyxl.Workbook()
        work_book.save(self.new_excel_name)
        work_book.close()

    # 控制class的表层run函数
    def spider_excel_run(self, begin_page_num, end_page_num):
        self.create_main_excel()
        sheet_names_all = self.control_range_home_url(begin_page_num, end_page_num)
        self.del_sheet1()
        return sheet_names_all


# 与上个class不同，此class用于生成不同省份随时间变化的数据，做到数据分类和整理，并存储到新的excel表中，并用于提供web的数据服务
class CoronaVirusWebmake(object):
    # 初始化方法，此处的sheet_names1为上一个表的sheet名列表，由于是用日期命名，直接用于日期列表
    def __init__(self, sheet_names1, excel_name1, excel_name2='new2'):
        self.sheet_days_in_data = sheet_names1  # 此处的名称为日期列表（上一个表中的sheet名）
        self.old_excel_name = excel_name1 + '.xlsx'  # 上一个excel名
        self.new_excel_name = excel_name2 + '.xlsx'  # 新建的excel表名
        self.main_sheet_name = ['本土新增确诊', '本土新增无症状', '港澳台累计确诊', '港澳台新增确诊']
        self.province_name = [
            '安徽', '北京', '重庆', '福建', '甘肃', '广东', '广西', '贵州', '海南', '河北', '河南',
            '黑龙江', '湖北', '湖南', '吉林', '江苏', '江西', '辽宁', '内蒙古', '宁夏', '青海', '山东',
            '山西', '陕西', '上海', '四川', '天津', '西藏', '新疆', '云南', '浙江'
        ]
        self.other_province_name = ['香港', '澳门', '台湾']

    # 新建一个excel表
    def create_province_excel(self):
        work_book = openpyxl.Workbook()
        for i in range(len(self.main_sheet_name)):
            work_book.create_sheet(self.main_sheet_name[i])

        sheet_name = 'Sheet'
        worksheet = work_book[sheet_name]
        work_book.remove(worksheet)

        work_book.save(self.new_excel_name)
        work_book.close()

    # 将一个省份的数据写入excel
    def write_one_province_excel(self, sheet_name, province_name, list_one_province, sheet_column):
        work_book = openpyxl.load_workbook(self.new_excel_name)
        sheet = work_book[sheet_name]
        sheet.cell(1, sheet_column, province_name)
        for i in range(len(list_one_province)):
            sheet.cell(i + 2, sheet_column, list_one_province[i])

        work_book.save(self.new_excel_name)
        work_book.close()

    # 核心控制，调用很多函数方法，生成需要的数据
    def control_write_excel(self):
        different_names = ['中国内地'] + self.province_name
        len_province = len(different_names)

        for k in range(2):
            sheet_name = self.main_sheet_name[k]
            self.insert_day_name(sheet_name)
            for i in tqdm(range(len_province), sheet_name + 'excel表生成中'):
                list_one_province = self.collect_one_province_data(i + 2, k + 2)
                self.write_one_province_excel(sheet_name, different_names[i], list_one_province, i + 2)

        for i in tqdm(range(3), self.main_sheet_name[2] + '和新增确诊excel表生成中'):
            self.insert_day_name(self.main_sheet_name[2])
            self.insert_day_name(self.main_sheet_name[3])
            list_one_province = self.collect_one_province_data(i + 35, 2)
            list_add = []
            for k in range(len(list_one_province) - 1):
                num_add = list_one_province[k] - list_one_province[k + 1]
                if num_add < 0:
                    num_add = 0
                list_add.append(num_add)
            list_add.append('...')
            self.write_one_province_excel(self.main_sheet_name[2], self.other_province_name[i], list_one_province,
                                          i + 2)
            self.write_one_province_excel(self.main_sheet_name[3], self.other_province_name[i], list_add, i + 2)
        # for i in range(len_province):
        #     list_one_province = self.collect_one_province_data(i + 2, 3)

    # 写入日期的数据
    def insert_day_name(self, sheet_name):
        work_book = openpyxl.load_workbook(self.new_excel_name)
        sheet = work_book[sheet_name]
        sheet.cell(1, 1, sheet_name)
        day_name = self.sheet_days_in_data
        for i in range(len(day_name)):
            sheet.cell(i + 2, 1, day_name[i])
        work_book.save(self.new_excel_name)
        work_book.close()

    # 收集一个省份的数据
    def collect_one_province_data(self, num_province, num_location):
        work_book = openpyxl.load_workbook(self.old_excel_name)
        day_all = self.sheet_days_in_data
        day_len = len(day_all)
        list_data = []
        for i in range(day_len):
            sheet = work_book[day_all[i]]
            list_data = list_data + [int(sheet.cell(num_province, num_location).value)]
        # print(list_data)
        return list_data

    # 修改表格单元格格式
    def change_style_excel(self):
        work_book = openpyxl.load_workbook(self.new_excel_name)
        for k in range(len(self.main_sheet_name)):
            sheet = work_book[self.main_sheet_name[k]]
            sheet.column_dimensions['A'].width = 15.0
            for i in range(sheet.max_column):
                for j in range(sheet.max_row):
                    sheet.cell(j + 1, i + 1).alignment = Alignment(horizontal='center', vertical='center')
        work_book.save(self.new_excel_name)

    # class的主要run函数
    def excel_run(self):
        self.create_province_excel()
        self.control_write_excel()
        self.change_style_excel()
        # list9 = self.collect_one_province_data(2, 2)
        # print(list9)

# 下面是先前的main函数，现在将这个写如main文件中去了，两个函数成功包装
# if __name__ == '__main__':
#     excel_name_1 = 'province'
#     excel_name_2 = 'province_in_diff_days'
#     begin_page = 1
#     end_page = 2
#
#     spider = CoronaVirusSpider(excel_name_1)
#     spider.spider_excel_run(begin_page, end_page)
# test = CoronaVirusWebmake(sheet_days_in_data, excel_name_1, excel_name_2)
# test.excel_run()
