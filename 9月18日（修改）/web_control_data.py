import re
import openpyxl


# 此class为控制web从对应的excel表中读取分好类的数据，返回给服务器flask函数
class GetDataFromExcel(object):
    # 初始化函数方法
    def __init__(self, excel_name_use):
        self.excel_name = excel_name_use + '.xlsx'
        self.sheet_names = self.return_sheet_names()
        self.province_list = self.return_province_names()
        self.province_dict = self.change_list_to_dict_func(self.province_list)
        self.other_province_name = ['香港', '澳门', '台湾']

    # 获取疫情省份名字
    def return_other_province_names(self):
        return self.other_province_name

    # 将当天省份疫情信息转为字典格式，主要用于中国地图的data格式转换
    def return_province_dict(self):
        return self.province_dict

    # 作为初始化获取excel省份名字的函数，如果有人将excel表做了数据排序等操作，此函数方法可以兼容保证省份对应位置正确（手打信息是不可靠的）
    def return_province_names(self):
        work_book = openpyxl.load_workbook(self.excel_name)
        sheet = work_book[self.sheet_names[0]]
        column = sheet.max_column + 1
        list_add = []
        for i in range(2, column):
            list_add.append(sheet.cell(1, i).value)
        work_book.close()
        return list_add

    # 作为初始化获取excel表的sheet名
    def return_sheet_names(self):
        work_book = openpyxl.load_workbook(self.excel_name)
        sheet_names = work_book.sheetnames
        work_book.close()
        return sheet_names

    # 改列表为字典
    @staticmethod
    def change_list_to_dict_func(list_1):
        dict_add = {}
        for i in range(len(list_1)):
            dict_add[list_1[i]] = i
        return dict_add

    # 改变日期格式从2022-01-01到1月1日
    @staticmethod
    def change_day_to_chinese(day_name):
        str_list = re.findall(r'-(\w+)', day_name)
        num1 = int(str_list[0])
        num2 = int(str_list[1])
        return str(num1) + '月' + str(num2) + '日'

    # 在中文日期格式前加某某年，如从1月1日到2022年1月1日
    @staticmethod
    def change_day_to_chinese_year(day_name):
        str_list1 = re.findall(r'(\w+)-', day_name)
        str_list2 = re.findall(r'-(\w+)', day_name)
        num1 = int(str_list1[0])
        num2 = int(str_list1[1])
        num3 = int(str_list2[1])
        return str(num1) + '年' + str(num2) + '月' + str(num3) + '日'

    # 获取表格所有信息，按行存入列表，再次强调是全部信息 按 行 ！
    def get_datas_all_collect_from_excel(self):
        work_book = openpyxl.load_workbook(self.excel_name)
        list_all = []
        for k in range(len(self.sheet_names)):
            sheet_name1 = self.sheet_names[k]
            sheet = work_book[sheet_name1]
            row = sheet.max_row
            column = sheet.max_column
            list1 = []
            for i in range(row):
                row_list = []
                for j in range(column):
                    row_list.append((sheet.cell(i + 1, j + 1)).value)
                list1.append(row_list)
            list_all.append(list1)
        work_book.close()
        return list_all

    # 获取几天内日期文本list(str),格式为1月1日,顺序从过去到现在，参数为天数(int)
    def get_days_for_excel(self, day_len, mold_num):
        work_book = openpyxl.load_workbook(self.excel_name)
        sheet = work_book[self.sheet_names[0]]
        list_days = []
        if mold_num == 0:
            for i in range(day_len):
                day_change = self.change_day_to_chinese(sheet.cell(day_len - i + 1, 1).value)
                list_days.append(day_change)
        else:
            for i in range(day_len):
                day_change = self.change_day_to_chinese_year(sheet.cell(day_len - i + 1, 1).value)
                list_days.append(day_change)
        work_book.close()
        return list_days

    # 获取几日内对应省份数据list(int)，顺序从过去到现在，参数1为省份名(str)，参数2为天数(int)
    def get_province_data(self, one_province, day_len):
        work_book = openpyxl.load_workbook(self.excel_name)
        if one_province in self.province_dict:
            province_num = self.province_dict[one_province]
        else:
            province_num = 0
        list_province_data = []
        for k in range(2):
            list_1 = []
            sheet = work_book[self.sheet_names[k]]
            for i in range(day_len):
                list_1.append(int(sheet.cell(day_len - i + 1, province_num + 2).value))
            list_province_data.append(list_1)
        work_book.close()
        return list_province_data

    # 获取港澳台省份几日内数据包list(int)，顺序从过去到现在，参数为天数(int)
    def get_others_province_data(self, day_len):
        work_book = openpyxl.load_workbook(self.excel_name)
        sheet = work_book[self.sheet_names[3]]
        list_others = []
        for k in range(3):
            list3 = []
            for i in range(day_len):
                list3.append(int(sheet.cell(day_len - i + 1, k + 2).value))
            list_others.append(list3)
        work_book.close()
        return list_others

    # 获取前某天的数据list(int)，默认省份排序，第一项为日期str：某月某日格式，参数为前(某)天(int)
    def get_one_day_data(self, day_num):
        work_book = openpyxl.load_workbook(self.excel_name)
        list_day = []
        for k in range(2):
            sheet = work_book[self.sheet_names[k]]
            list_temp = []
            for i in range(2, 34):
                list_temp.append(sheet.cell(day_num + 2, i).value)
            list_day.append(list_temp)
        return list_day

    # def test(self, day_num):
    #     work_book = openpyxl.load_workbook(self.excel_name)
    #     list_day = []
    #     for k in range(2):
    #         sheet = work_book[self.sheet_names[k]]
    #         list_temp = []
    #         for i in range(2, 34):
    #             list_temp.append(sheet.cell(day_num + 2, i).value)
    #         list_day.append(list_temp)
    #     return list_day

    # 用于生成中国地图信息列表，第一个大列表为本土新增确诊，第二个为新增无症状，第三个为港澳台新增（地图显示在新增确诊里）
    def get_map_data(self, one_day):
        list_data = self.get_one_day_data(one_day)
        pro_len = len(list_data[0])
        list_return = []
        for k in range(2):
            list_list = []
            for i in range(1, pro_len):
                list_list.append(list_data[k][i])
            list_return.append(list_list)
        list_others = self.get_others_province_data(one_day + 1)
        list_other = [list_others[0][0], list_others[1][0], list_others[2][0]]
        list_return.append(list_other)
        return list_return

    # 寻找近day_num天新增确诊病例的热点信息，很简单的分析，只找出最大差值再缩小区间，这部分的能力很弱
    def get_hot_point(self, day_num):
        work_book = openpyxl.load_workbook(self.excel_name)
        sheet = work_book[self.sheet_names[0]]
        province_name = self.province_list
        day_name = self.get_days_for_excel(day_num, 0)
        day_len = len(day_name)
        flag_num = [0, 0, 0]
        max_province = 0
        max_num_province = [0]
        for k in range(3, 32):
            min_num = 2
            max_num = 2
            for i in range(3, day_len + 2):
                temp = sheet.cell(i, k).value
                if temp > sheet.cell(max_num, k).value:
                    max_num = i
                if temp < sheet.cell(min_num, k).value:
                    min_num = i
            max_add = sheet.cell(max_num, k).value - sheet.cell(min_num, k).value
            if max_province < max_add:
                max_num_province[0] = k
                max_province = max_add
                if min_num < max_num:
                    for i in range(min_num + 1, max_num):
                        if sheet.cell(i, k).value == sheet.cell(min_num, k).value:
                            min_num = i
                        else:
                            break
                    flag_num[2] = 1
                    flag_num[0] = min_num
                    flag_num[1] = max_num
                else:
                    for i in range(max_num + 1, min_num):
                        if sheet.cell(i, k).value == sheet.cell(max_num, k).value:
                            max_num = i
                        else:
                            break
                    flag_num[2] = 0
                    flag_num[1] = min_num
                    flag_num[0] = max_num
        # print(day_name)
        # print(flag_num)
        # print(type(str(max_num_province[1])))
        if flag_num[2] == 0:
            list_hot = province_name[max_num_province[0] - 2] + '从' + day_name[day_len - flag_num[1] + 1] + '到' + \
                       day_name[day_len - flag_num[0] + 1] + '增加了' + str(max_province) + '例'
        else:
            list_hot = province_name[max_num_province[0] - 2] + '从' + day_name[day_len - flag_num[1] + 1] + '到' + \
                       day_name[day_len - flag_num[0] + 1] + '减少了' + str(max_province) + '例'

        list_return = ["/province/" + province_name[max_num_province[0] - 2], list_hot]
        return list_return

# if __name__ == '__main__':
#     test = GetDataFromExcel('province_in_diff_days')
#     print(test.get_one_day_data(7))
#     print(test.test(7))
